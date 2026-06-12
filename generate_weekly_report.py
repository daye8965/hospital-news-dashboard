"""
주간언론실적 자동 생성 스크립트
A. 기존 양식 그대로 엑셀 생성
B. 제약회사 홍보기사 고도화 필터
C. 집계표 HTML 자동 생성
"""
import os, csv, re, html, sys
from datetime import datetime, timedelta, timezone, date
from email.utils import parsedate_to_datetime
from pathlib import Path
import requests
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                               GradientFill)
from openpyxl.utils import get_column_letter

# ── API 키 ────────────────────────────────────────────────────────────────────
CLIENT_ID     = os.environ.get("NAVER_CLIENT_ID",     "_eBIxCwm7W7VId2GzO2R")
CLIENT_SECRET = os.environ.get("NAVER_CLIENT_SECRET", "itD8WVlnrL")

# ── 설정 ─────────────────────────────────────────────────────────────────────
KST = timezone(timedelta(hours=9))

# 이번 주 월~일 자동 계산 (실행 시점 기준)
def get_week_range():
    today = datetime.now(KST).date()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)
    return monday, sunday

WEEK_START, WEEK_END = get_week_range()
# CLI 인자로 날짜 지정 가능: python script.py 2026-06-01 2026-06-07
if len(sys.argv) == 3:
    WEEK_START = date.fromisoformat(sys.argv[1])
    WEEK_END   = date.fromisoformat(sys.argv[2])

start_dt = datetime(WEEK_START.year, WEEK_START.month, WEEK_START.day, tzinfo=KST)
end_dt   = datetime(WEEK_END.year,   WEEK_END.month,   WEEK_END.day,   23, 59, 59, tzinfo=KST)

HOSPITAL_QUERIES = {
    "서울아산병원": ["서울아산병원", "아산병원", "울산의대", "서울중앙병원"],
    "삼성서울병원": ["삼성서울병원", "삼성병원", "성대의대", "성균관의대"],
    "신촌세브란스": ["세브란스병원", "연세암병원", "연대병원", "연세대병원"],
    "서울대병원":   ["서울대병원", "서울대학교병원", "서울의대"],
    "서울성모병원": ["서울성모병원", "카톨릭성모병원"],
}
HOSPITAL_KEYS = list(HOSPITAL_QUERIES.keys())

# ── B. 고도화 제외 필터 ───────────────────────────────────────────────────────
EXCLUDE_EXACT   = {"부고", "부음", "별세", "추모", "조문"}
EXCLUDE_PATTERN = re.compile(
    r"(부고|부음|별세|타계|영면|빈소|발인|장례|추도|추모식"
    r"|보도자료|출시|신제품|식약처\s*허가|FDA\s*(승인|허가)"
    r"|임상\s*[0-9]?상\s*승인|글로벌\s*출시|국내\s*출시"
    r"|MOU\s*체결|업무협약|투자\s*유치|시리즈\s*[A-Z]"
    r"|제약|바이오텍|헬스케어\s*기업|상장|주가|매출|영업이익)"
)

BROADCAST = {"KBS", "MBC", "SBS", "EBS", "JTBC", "YTN", "MBN", "TV조선",
             "채널A", "연합뉴스TV", "OBS", "KNN", "TBC", "MBC경남"}
DAILY_PRESS = {"조선일보","중앙일보","동아일보","매일경제","한국경제",
               "한겨레","경향신문","한국일보","서울신문","국민일보",
               "세계일보","문화일보","서울경제","파이낸셜뉴스","머니투데이",
               "헤럴드경제","아시아경제","이데일리","전자신문"}

MEDIA_MAP = {
    "chosun.com":"조선일보","joongang.co.kr":"중앙일보","donga.com":"동아일보",
    "hani.co.kr":"한겨레","khan.co.kr":"경향신문","hankyung.com":"한국경제",
    "mk.co.kr":"매일경제","sedaily.com":"서울경제","seoul.co.kr":"서울신문",
    "kookmin.co.kr":"국민일보","segye.com":"세계일보","munhwa.com":"문화일보",
    "hankookilbo.com":"한국일보","fnnews.com":"파이낸셜뉴스","mt.co.kr":"머니투데이",
    "heraldcorp.com":"헤럴드경제","asiae.co.kr":"아시아경제","edaily.co.kr":"이데일리",
    "etnews.com":"전자신문","yna.co.kr":"연합뉴스","newsis.com":"뉴시스",
    "news1.kr":"뉴스1","kbs.co.kr":"KBS","imbc.com":"MBC","sbs.co.kr":"SBS",
    "ebs.co.kr":"EBS","jtbc.co.kr":"JTBC","ytn.co.kr":"YTN","mbn.co.kr":"MBN",
    "tvchosun.com":"TV조선","ichannela.com":"채널A","healthchosun.com":"헬스조선",
    "kormedi.com":"코메디닷컴","medicaltimes.com":"메디칼타임즈",
    "doctorsnews.co.kr":"의사신문","rapportian.com":"라포르시안",
    "mdtoday.co.kr":"메디컬투데이","mdjournal.kr":"메디컬저널",
    "bosa.co.kr":"보건복지부","newsmp.com":"메디컬포스트",
    "dailymedi.com":"데일리메디","pharmnews.com":"팜뉴스",
    "yakup.com":"약업신문","kpanews.co.kr":"약사공론",
}

def extract_media(url):
    if not url:
        return ""
    try:
        domain = re.sub(r"https?://(www\.)?", "", url).split("/")[0].lower()
        for key, name in MEDIA_MAP.items():
            if key in domain:
                return name
        parts = domain.split(".")
        return parts[-2] if len(parts) >= 2 else domain
    except:
        return ""

def clean_text(text):
    return html.unescape(re.sub(r"<.*?>", "", text)).strip()

def is_excluded(title, summary=""):
    combined = title + " " + summary
    if any(kw in title for kw in EXCLUDE_EXACT):
        return True
    if EXCLUDE_PATTERN.search(combined):
        return True
    return False

def media_type(media):
    if media in BROADCAST:     return "방송"
    if media in DAILY_PRESS:   return "신문"
    return "온라인"

# ── 네이버 수집 ───────────────────────────────────────────────────────────────
def fetch_news(query):
    collected, seen = [], set()
    for start in range(1, 1001, 100):
        try:
            resp = requests.get(
                "https://openapi.naver.com/v1/search/news.json",
                headers={"X-Naver-Client-Id": CLIENT_ID,
                         "X-Naver-Client-Secret": CLIENT_SECRET},
                params={"query": query, "display": 100, "start": start, "sort": "date"},
                timeout=10,
            )
            resp.raise_for_status()
        except Exception as e:
            print(f"  API 오류 [{query}]: {e}")
            break
        items = resp.json().get("items", [])
        if not items: break
        for item in items:
            pub_dt = parsedate_to_datetime(item["pubDate"]).astimezone(KST)
            if not (start_dt <= pub_dt <= end_dt): continue
            title   = clean_text(item.get("title",""))
            summary = clean_text(item.get("description",""))
            if is_excluded(title, summary): continue
            orig = item.get("originallink","")
            key  = orig or item.get("link","")
            if key in seen: continue
            seen.add(key)
            media = extract_media(orig)
            collected.append({
                "날짜":    pub_dt.strftime("%Y-%m-%d"),
                "매체":    media,
                "유형":    media_type(media),
                "제목":    title,
                "링크":    orig or item.get("link",""),
                "진료과":  "",
                "이름":    "",
                "비고":    "",
            })
        last = parsedate_to_datetime(items[-1]["pubDate"]).astimezone(KST)
        if last < start_dt: break
    return collected

def collect_all():
    result = {}
    for hosp, queries in HOSPITAL_QUERIES.items():
        seen_keys, rows = set(), []
        for q in queries:
            news = fetch_news(q)
            for n in news:
                k = n["링크"] or n["제목"]
                if k not in seen_keys:
                    seen_keys.add(k)
                    rows.append(n)
            print(f"  [{hosp}] {q}: {len(news)}건")
        rows.sort(key=lambda x: x["날짜"])
        result[hosp] = rows
        print(f"  ▶ [{hosp}] 합계: {len(rows)}건\n")
    return result

# ── 스타일 헬퍼 ───────────────────────────────────────────────────────────────
def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)

def header_fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def cell(ws, row, col, value="", bold=False, center=False,
         fill=None, font_color="000000", size=10, border=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name="맑은 고딕", size=size, bold=bold, color=font_color)
    c.alignment = Alignment(
        horizontal="center" if center else "left",
        vertical="center", wrap_text=wrap)
    if fill: c.fill = fill
    if border: c.border = thin_border()
    return c

# ── A. 주간실적 엑셀 생성 ─────────────────────────────────────────────────────
def make_summary_sheet(wb, data, week_label):
    ws = wb.create_sheet("요약", 0)
    ws.sheet_view.showGridLines = False

    # 제목
    ws.merge_cells("B2:G2")
    c = ws["B2"]
    c.value = "주요병원 주간언론실적"
    c.font = Font(name="맑은 고딕", size=14, bold=True)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("B3:G3")
    c = ws["B3"]
    c.value = week_label
    c.font = Font(name="맑은 고딕", size=10)
    c.alignment = Alignment(horizontal="center", vertical="center")

    ws.merge_cells("G4:G4")
    ws["H4"] = "단위: 건"
    ws["H4"].font = Font(name="맑은 고딕", size=9, color="666666")
    ws["H4"].alignment = Alignment(horizontal="right")

    # 헤더
    HOSP_COL = {h: i+4 for i, h in enumerate(HOSPITAL_KEYS)}
    header_bg = header_fill("1F4E79")
    sub_bg    = header_fill("2E75B6")
    row_bg    = header_fill("DEEAF1")

    ws.merge_cells("B5:C5")
    cell(ws, 5, 2, "구분", bold=True, center=True, fill=header_bg, font_color="FFFFFF")
    for hosp, col in HOSP_COL.items():
        cell(ws, 5, col, hosp, bold=True, center=True, fill=header_bg, font_color="FFFFFF")

    # 신문 행
    PAPERS = ["조선일보","중앙일보","동아일보","매일경제","한국경제","기타","계 >"]
    BROADCASTS = ["KBS","SBS","MBC","기타","계 >"]

    def count_media(rows, media_name, mtype="신문"):
        if media_name == "기타":
            if mtype == "신문":
                named = {"조선일보","중앙일보","동아일보","매일경제","한국경제"}
                return sum(1 for r in rows if r["유형"]=="신문" and r["매체"] not in named)
            else:
                named = {"KBS","SBS","MBC"}
                return sum(1 for r in rows if r["유형"]=="방송" and r["매체"] not in named)
        if media_name == "계 >":
            return sum(1 for r in rows if r["유형"]==mtype)
        return sum(1 for r in rows if r["매체"]==media_name)

    # 신문 섹션
    r = 6
    ws.merge_cells(f"B{r}:B{r+len(PAPERS)-1}")
    c = ws.cell(row=r, column=2, value="신문")
    c.font = Font(name="맑은 고딕", bold=True, size=10)
    c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.fill = sub_bg; c.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    c.border = thin_border()

    for i, paper in enumerate(PAPERS):
        is_total = paper == "계 >"
        bg = header_fill("BDD7EE") if is_total else (row_bg if i%2==0 else None)
        cell(ws, r+i, 3, paper, bold=is_total, center=True, fill=bg)
        for hosp, col in HOSP_COL.items():
            rows = data.get(hosp, [])
            mtype = "신문"
            v = count_media(rows, paper, mtype)
            cell(ws, r+i, col, v if v else 0, center=True,
                 fill=bg, bold=is_total)

    # 방송 섹션
    r2 = r + len(PAPERS)
    ws.merge_cells(f"B{r2}:B{r2+len(BROADCASTS)-1}")
    c = ws.cell(row=r2, column=2, value="방송\n(뉴스)")
    c.font = Font(name="맑은 고딕", bold=True, color="FFFFFF", size=10)
    c.fill = sub_bg; c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    c.border = thin_border()

    for i, bc in enumerate(BROADCASTS):
        is_total = bc == "계 >"
        bg = header_fill("BDD7EE") if is_total else (row_bg if i%2==0 else None)
        cell(ws, r2+i, 3, bc, bold=is_total, center=True, fill=bg)
        for hosp, col in HOSP_COL.items():
            rows = data.get(hosp, [])
            v = count_media(rows, bc, "방송")
            cell(ws, r2+i, col, v if v else 0, center=True,
                 fill=bg, bold=is_total)

    # 합계 행
    r3 = r2 + len(BROADCASTS)
    total_bg = header_fill("1F4E79")
    ws.merge_cells(f"B{r3}:C{r3}")
    cell(ws, r3, 2, "신문+방송 합계 >", bold=True, center=True,
         fill=total_bg, font_color="FFFFFF")
    for hosp, col in HOSP_COL.items():
        rows = data.get(hosp, [])
        v = sum(1 for row in rows if row["유형"] in ("신문","방송"))
        cell(ws, r3, col, v, bold=True, center=True,
             fill=total_bg, font_color="FFFFFF")

    # 온라인 기사 수 주석
    asan_online = sum(1 for r in data.get("서울아산병원",[]) if r["유형"]=="온라인")
    ws.cell(row=r3+1, column=2).value = f"* 서울아산병원 온라인 기사: {asan_online}건"
    ws.cell(row=r3+1, column=2).font = Font(name="맑은 고딕", size=9, color="666666")

    # 컬럼 너비
    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 12
    for hosp, col in HOSP_COL.items():
        ws.column_dimensions[get_column_letter(col)].width = 14
    for i in range(6, r3+2):
        ws.row_dimensions[i].height = 18
    ws.row_dimensions[2].height = 28
    ws.row_dimensions[5].height = 22


def make_hospital_sheet(wb, hosp_name, rows):
    ws = wb.create_sheet(hosp_name)
    ws.sheet_view.showGridLines = False

    ws.column_dimensions["A"].width = 2
    ws.column_dimensions["B"].width = 6
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 52
    ws.column_dimensions["F"].width = 16
    ws.column_dimensions["G"].width = 8
    ws.column_dimensions["H"].width = 8

    title_bg = header_fill("1F4E79")
    section_bg = header_fill("2E75B6")
    header_bg2 = header_fill("DEEAF1")

    # 제목
    ws.merge_cells("B2:H2")
    c = ws["B2"]
    c.value = f"{hosp_name} 언론 실적"
    c.font = Font(name="맑은 고딕", size=12, bold=True, color="FFFFFF")
    c.fill = title_bg
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 24

    def section(ws, start_row, label, section_rows):
        # 섹션 헤더
        ws.merge_cells(f"B{start_row}:H{start_row}")
        c = ws.cell(row=start_row, column=2, value=f"• {label}")
        c.font = Font(name="맑은 고딕", size=10, bold=True, color="FFFFFF")
        c.fill = section_bg
        c.alignment = Alignment(horizontal="left", vertical="center")
        c.border = thin_border()
        ws.row_dimensions[start_row].height = 20

        # 컬럼 헤더
        h = start_row + 1
        headers = ["순 서","날 짜","매 체","보  도  내  용","진 료 과","이 름","비 고"]
        for i, hdr in enumerate(headers):
            cell(ws, h, i+2, hdr, bold=True, center=True, fill=header_bg2)
        ws.row_dimensions[h].height = 18

        # 데이터
        for j, row in enumerate(section_rows):
            r = h + 1 + j
            alt = header_fill("F5F9FF") if j % 2 == 0 else None
            cell(ws, r, 2, j+1, center=True, fill=alt)
            cell(ws, r, 3, row["날짜"], center=True, fill=alt)
            cell(ws, r, 4, row["매체"], center=True, fill=alt)
            cell(ws, r, 5, row["제목"], fill=alt, wrap=True)
            cell(ws, r, 6, row.get("진료과",""), center=True, fill=alt)
            cell(ws, r, 7, row.get("이름",""), center=True, fill=alt)
            cell(ws, r, 8, row.get("비고",""), center=True, fill=alt)
            ws.row_dimensions[r].height = 18
        return h + 1 + len(section_rows)

    papers    = [r for r in rows if r["유형"] == "신문"]
    broadcasts= [r for r in rows if r["유형"] == "방송"]
    online    = [r for r in rows if r["유형"] == "온라인"]

    next_row = section(ws, 4, "신문", papers) + 1
    next_row = section(ws, next_row, "방송(뉴스)", broadcasts) + 1
    section(ws, next_row, "온라인 기사", online)


# ── C. 집계표 HTML 생성 ───────────────────────────────────────────────────────
def make_summary_html(data, week_label, out_path):
    PAPERS = ["조선일보","중앙일보","동아일보","매일경제","한국경제","기타"]
    BCAST  = ["KBS","SBS","MBC","기타"]

    def cnt(rows, name, mtype):
        if name == "기타":
            named = ({"조선일보","중앙일보","동아일보","매일경제","한국경제"}
                     if mtype=="신문" else {"KBS","SBS","MBC"})
            return sum(1 for r in rows if r["유형"]==mtype and r["매체"] not in named)
        return sum(1 for r in rows if r["매체"]==name)

    def total(rows, mtype):
        return sum(1 for r in rows if r["유형"]==mtype)

    hospitals = HOSPITAL_KEYS
    col_w = 100

    def td(val, bold=False, bg="", color="#1a1a18"):
        s = f"font-weight:{'700' if bold else '400'};"
        if bg: s += f"background:{bg};"
        if color != "#1a1a18": s += f"color:{color};"
        return f'<td style="{s}text-align:center;padding:7px 10px;border:1px solid #dee2e6">{val if val != 0 or not bold else val}</td>'

    rows_html = ""
    for i, paper in enumerate(PAPERS):
        bg = "#f8f9fa" if i%2==0 else "#ffffff"
        rows_html += f'<tr style="background:{bg}"><td style="padding:7px 14px;border:1px solid #dee2e6">{paper}</td>'
        for h in hospitals:
            v = cnt(data.get(h,[]), paper, "신문")
            rows_html += td(v or "—")
        rows_html += "</tr>\n"

    # 신문 계
    rows_html += '<tr style="background:#dbeafe"><td style="padding:7px 14px;border:1px solid #dee2e6;font-weight:700">계 ›</td>'
    paper_totals = [total(data.get(h,[]), "신문") for h in hospitals]
    for v in paper_totals:
        rows_html += td(v, bold=True, bg="#dbeafe", color="#1e40af")
    rows_html += "</tr>\n"

    for i, bc in enumerate(BCAST):
        bg = "#f8f9fa" if i%2==0 else "#ffffff"
        rows_html += f'<tr style="background:{bg}"><td style="padding:7px 14px;border:1px solid #dee2e6">{bc}</td>'
        for h in hospitals:
            v = cnt(data.get(h,[]), bc, "방송")
            rows_html += td(v or "—")
        rows_html += "</tr>\n"

    # 방송 계
    rows_html += '<tr style="background:#dbeafe"><td style="padding:7px 14px;border:1px solid #dee2e6;font-weight:700">계 ›</td>'
    bc_totals = [total(data.get(h,[]), "방송") for h in hospitals]
    for v in bc_totals:
        rows_html += td(v, bold=True, bg="#dbeafe", color="#1e40af")
    rows_html += "</tr>\n"

    # 합계
    rows_html += '<tr style="background:#1e3a5f"><td style="padding:9px 14px;border:1px solid #1e3a5f;font-weight:700;color:#fff">신문+방송 합계 ›</td>'
    for h in hospitals:
        v = paper_totals[hospitals.index(h)] + bc_totals[hospitals.index(h)]
        rows_html += f'<td style="text-align:center;padding:9px 10px;border:1px solid #1e3a5f;font-weight:700;color:#fff;background:#1e3a5f">{v}</td>'
    rows_html += "</tr>\n"

    asan_online = sum(1 for r in data.get("서울아산병원",[]) if r["유형"]=="온라인")

    header_cols = "".join(
        f'<th style="background:#1e3a5f;color:#fff;padding:10px;border:1px solid #1e3a5f;min-width:{col_w}px">{h}</th>'
        for h in hospitals
    )

    html_content = f"""<!DOCTYPE html>
<html lang="ko">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>주요병원 주간언론실적 {week_label}</title>
<link href="https://fonts.googleapis.com/css2?family=Noto+Sans+KR:wght@400;500;700&display=swap" rel="stylesheet">
<style>
  *{{box-sizing:border-box;margin:0;padding:0}}
  body{{font-family:'Noto Sans KR',sans-serif;background:#f4f6f9;min-height:100vh;display:flex;align-items:flex-start;justify-content:center;padding:40px 20px}}
  .card{{background:#fff;border-radius:12px;box-shadow:0 4px 24px rgba(0,0,0,.08);padding:36px 40px;max-width:900px;width:100%}}
  .title{{font-size:20px;font-weight:700;text-align:center;color:#1a1a18;margin-bottom:4px}}
  .subtitle{{font-size:13px;color:#6b6b67;text-align:center;margin-bottom:6px}}
  .unit{{font-size:11px;color:#9b9b97;text-align:right;margin-bottom:10px}}
  table{{width:100%;border-collapse:collapse;font-size:13px}}
  .section-label{{background:#2e75b6;color:#fff;font-weight:700;padding:8px 14px;border:1px solid #2e75b6;font-size:12px;letter-spacing:.5px}}
  .note{{font-size:11px;color:#6b6b67;margin-top:10px;text-align:right}}
  .btn{{display:inline-flex;align-items:center;gap:6px;margin-top:20px;padding:8px 18px;background:#1e3a5f;color:#fff;border:none;border-radius:6px;font-family:inherit;font-size:13px;cursor:pointer;text-decoration:none}}
  .btn:hover{{background:#2e5080}}
  @media print{{body{{background:#fff;padding:0}} .card{{box-shadow:none;padding:20px}} .btn{{display:none}}}}
</style>
</head>
<body>
<div class="card">
  <div class="title">주요병원 주간언론실적</div>
  <div class="subtitle">{week_label}</div>
  <div class="unit">단위: 건</div>
  <table>
    <thead>
      <tr>
        <th style="background:#1e3a5f;color:#fff;padding:10px;border:1px solid #1e3a5f;min-width:80px">구분</th>
        {header_cols}
      </tr>
    </thead>
    <tbody>
      <tr><td class="section-label" colspan="{len(hospitals)+1}">신문</td></tr>
      {rows_html}
    </tbody>
  </table>
  <div class="note">* 서울아산병원 온라인 기사: {asan_online}건</div>
  <div style="text-align:center">
    <button class="btn" onclick="window.print()">🖨 인쇄 / 이미지 저장</button>
  </div>
</div>
</body>
</html>"""

    Path(out_path).write_text(html_content, encoding="utf-8")
    print(f"HTML 저장: {out_path}")


# ── 실행 ─────────────────────────────────────────────────────────────────────
if __name__ == "__main__":
    week_label = (f"{WEEK_START.year}.{WEEK_START.month}.{WEEK_START.day}"
                  f"({['월','화','수','목','금','토','일'][WEEK_START.weekday()]}) ~ "
                  f"{WEEK_END.year}.{WEEK_END.month}.{WEEK_END.day}"
                  f"({['월','화','수','목','금','토','일'][WEEK_END.weekday()]})")

    print(f"\n📰 주간언론실적 자동 생성")
    print(f"대상 기간: {week_label}\n")

    # 수집
    data = collect_all()
    total = sum(len(v) for v in data.values())
    print(f"\n총 {total}건 수집 완료\n")

    # A. 엑셀 생성
    wb = Workbook()
    wb.remove(wb.active)
    make_summary_sheet(wb, data, week_label)
    for hosp in HOSPITAL_KEYS:
        make_hospital_sheet(wb, hosp, data.get(hosp, []))

    fname = WEEK_START.strftime("%Y%m%d")
    xlsx_path = f"/mnt/user-data/outputs/주간언론실적_{fname}.xlsx"
    wb.save(xlsx_path)
    print(f"✅ 엑셀 저장: {xlsx_path}")

    # C. HTML 집계표 생성
    html_path = f"/mnt/user-data/outputs/주간언론실적_{fname}.html"
    make_summary_html(data, week_label, html_path)
    print(f"✅ HTML 저장: {html_path}")
    print("\n🎉 완료!")
